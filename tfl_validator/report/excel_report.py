"""Excel Validation Report Generator.
Creates a professional workbook with summary dashboard, per-TFL detail sheets,
audit log, config documentation, and ADaM specs reference.
"""
import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


NAVY = "1F3864"
BLUE = "2E75B6"
LBLUE = "BDD7EE"
GREEN = "C6EFCE"
GREEN_F = "375623"
RED = "FFC7CE"
RED_F = "9C0006"
YELLOW = "FFF2CC"
GREY = "F2F2F2"
WHITE = "FFFFFF"
ORANGE = "FCE4D6"

def _hdr(cell, text, bg=NAVY, sz=10):
    cell.value = text
    cell.font = Font(name="Arial", size=sz, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(*[Side(style="thin", color="BFBFBF")]*4)

def _cell(cell, val, bg=WHITE, bold=False, color="000000", align="left"):
    cell.value = val
    cell.font = Font(name="Arial", size=10, bold=bold, color=color)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = Border(*[Side(style="thin", color="BFBFBF")]*4)


def generate_report(validation_results, audit_logger, config, output_path):
    """Generate the full validation Excel report.
    Args:
        validation_results: list of ValidationResult objects
        audit_logger: AuditLogger instance
        config: dict with study info and TFL configs
        output_path: path to save the .xlsx file
    """
    wb = Workbook()

    # ── SUMMARY SHEET ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "SUMMARY"
    ws.sheet_view.showGridLines = False

    study = config.get("study_info", {})
    ws.merge_cells("A1:H1")
    _hdr(ws["A1"], "TFL VALIDATION REPORT — SUMMARY", sz=14)
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    info = f"Study: {study.get('study_id','')} | {study.get('analysis','')} | Generated: {datetime.datetime.now().strftime('%d-%b-%Y %H:%M')}"
    _cell(ws["A2"], info, bg=LBLUE, bold=True, color=NAVY, align="center")

    # Summary stats
    total = len(validation_results)
    passed = sum(1 for v in validation_results if v.passed)
    failed = total - passed
    total_checks = sum(v.total for v in validation_results)
    total_pass = sum(v.pass_count for v in validation_results)
    total_fail = sum(v.fail_count for v in validation_results)

    r = 4
    for label, val, bg in [
        ("Total TFLs Validated", total, LBLUE),
        ("TFLs PASSED", passed, GREEN),
        ("TFLs FAILED", failed, RED if failed > 0 else GREEN),
        ("Total Checks Performed", total_checks, LBLUE),
        ("Individual Checks Passed", total_pass, GREEN),
        ("Individual Checks Failed", total_fail, RED if total_fail > 0 else GREEN),
    ]:
        _cell(ws.cell(r, 1), label, bg=GREY, bold=True, color=NAVY)
        _cell(ws.cell(r, 2), val, bg=bg, bold=True, color=RED_F if "FAIL" in label and val > 0 else GREEN_F if "PASS" in label else NAVY, align="center")
        ws.merge_cells(f"A{r}:A{r}")
        r += 1

    # TFL summary table
    r += 1
    headers = ["TFL ID", "Title", "Status", "Total Checks", "Passed", "Failed", "Match Rate", "Critical Findings"]
    for i, h in enumerate(headers):
        _hdr(ws.cell(r, i+1), h)
    ws.row_dimensions[r].height = 28

    for vr in validation_results:
        r += 1
        s = vr.summary()
        is_pass = s["status"] == "PASS"
        status_bg = GREEN if is_pass else RED
        status_color = GREEN_F if is_pass else RED_F

        vals = [s["tfl_id"], s["tfl_title"], s["status"], s["total_checks"],
                s["passed"], s["failed"], s["match_rate"],
                "None" if is_pass else f"{s['failed']} mismatch(es) found"]

        for i, v in enumerate(vals):
            bg = WHITE
            c = "000000"
            b = False
            if i == 2:
                bg = status_bg; c = status_color; b = True
            elif i == 5 and not is_pass:
                bg = RED; c = RED_F; b = True
            _cell(ws.cell(r, i+1), v, bg=bg, bold=b, color=c, align="center" if i >= 2 else "left")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 50
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 16

    # ── PER-TFL DETAIL SHEETS ────────────────────────────────────────────────
    for vr in validation_results:
        sheet_name = vr.tfl_id.replace("/", "_")[:31]
        ws_d = wb.create_sheet(sheet_name)
        ws_d.sheet_view.showGridLines = False

        ws_d.merge_cells("A1:G1")
        _hdr(ws_d["A1"], f"{vr.tfl_id} — VALIDATION DETAIL", sz=12)
        ws_d.row_dimensions[1].height = 30

        ws_d.merge_cells("A2:G2")
        _cell(ws_d["A2"], vr.tfl_title, bg=LBLUE, bold=True, color=NAVY, align="center")

        # Status banner
        status_bg = GREEN if vr.passed else RED
        status_color = GREEN_F if vr.passed else RED_F
        ws_d.merge_cells("A3:G3")
        _cell(ws_d["A3"],
              f"{'✓ PASSED' if vr.passed else '✗ FAILED'} — {vr.pass_count}/{vr.total} checks passed ({vr.match_rate:.1%})",
              bg=status_bg, bold=True, color=status_color, align="center")

        # Detail table
        r = 5
        detail_headers = ["#", "Statistic / Check", "Row Context", "TFL Value", "Recalculated Value", "Result", "Note"]
        for i, h in enumerate(detail_headers):
            _hdr(ws_d.cell(r, i+1), h)
        ws_d.row_dimensions[r].height = 26

        for idx, comp in enumerate(vr.comparisons, 1):
            r += 1
            match = comp["match"]
            row_bg = WHITE if match else "FFF0F0"
            result_bg = GREEN if match else RED
            result_color = GREEN_F if match else RED_F

            _cell(ws_d.cell(r, 1), idx, bg=GREY, align="center")
            _cell(ws_d.cell(r, 2), comp["stat_name"], bg=row_bg, bold=True, color=NAVY)
            _cell(ws_d.cell(r, 3), comp["row_label"] or comp.get("col_label", ""), bg=row_bg)
            _cell(ws_d.cell(r, 4), comp["tfl_value"], bg=row_bg, align="center")
            _cell(ws_d.cell(r, 5), comp["calc_value"], bg=row_bg, align="center")
            _cell(ws_d.cell(r, 6), "PASS" if match else "FAIL", bg=result_bg, bold=True, color=result_color, align="center")
            _cell(ws_d.cell(r, 7), comp["note"], bg=row_bg)

        ws_d.column_dimensions["A"].width = 6
        ws_d.column_dimensions["B"].width = 32
        ws_d.column_dimensions["C"].width = 26
        ws_d.column_dimensions["D"].width = 18
        ws_d.column_dimensions["E"].width = 20
        ws_d.column_dimensions["F"].width = 10
        ws_d.column_dimensions["G"].width = 44

        ws_d.sheet_properties.tabColor = GREEN_F if vr.passed else RED_F

    # ── AUDIT LOG SHEET ──────────────────────────────────────────────────────
    ws_a = wb.create_sheet("AUDIT LOG")
    ws_a.sheet_view.showGridLines = False

    ws_a.merge_cells("A1:L1")
    _hdr(ws_a["A1"], "CALCULATION AUDIT LOG — All computations logged for regulatory traceability", sz=11)
    ws_a.row_dimensions[1].height = 30

    ws_a.merge_cells("A2:L2")
    _cell(ws_a["A2"],
          "⚠ This log records every statistical calculation performed during validation. Each entry references the exact source file, function, and line number. Do not modify.",
          bg="FFF0F0", bold=True, color=RED_F, align="center")

    audit_headers = ["#", "Timestamp", "TFL ID", "Statistic", "Variable",
                     "Source File", "Function", "Line #", "Line Range",
                     "Code / Methodology", "Input Summary", "Result"]
    r = 4
    for i, h in enumerate(audit_headers):
        bg = NAVY if i < 5 else ("4472C4" if i < 9 else NAVY)
        _hdr(ws_a.cell(r, i+1), h, bg=bg)

    # Section label row
    r_label = 3
    ws_a.merge_cells(f"A{r_label}:E{r_label}")
    _cell(ws_a.cell(r_label, 1), "◄ WHAT was calculated", bg=NAVY, bold=True, color="FFFFFF", align="center")
    ws_a.merge_cells(f"F{r_label}:I{r_label}")
    _cell(ws_a.cell(r_label, 6), "◄ WHERE in source code (for audit traceability)", bg="4472C4", bold=True, color="FFFFFF", align="center")
    ws_a.merge_cells(f"J{r_label}:L{r_label}")
    _cell(ws_a.cell(r_label, 10), "◄ HOW and RESULT", bg=NAVY, bold=True, color="FFFFFF", align="center")

    for entry in audit_logger.get_entries():
        r += 1
        vals = [
            entry["entry_num"],
            entry["timestamp"],
            entry["tfl_id"],
            entry["statistic"],
            entry["variable"],
            entry.get("source_file", ""),
            entry.get("function", ""),
            entry.get("line_number", ""),
            entry.get("line_range", ""),
            entry["code_description"],
            entry["input_summary"],
            entry["result"],
        ]
        is_fail = str(entry["result"]).startswith("FAIL")
        is_compare = entry["code_description"].startswith("COMPARE")
        for i, v in enumerate(vals):
            bg = WHITE
            if is_fail: bg = "FFF0F0"
            elif is_compare and str(entry["result"]).startswith("PASS"): bg = "F0FFF0"
            # Source code columns get a light blue tint
            if 5 <= i <= 8 and not is_fail:
                bg = "E8F0FE"
            c = "000000"
            if is_fail: c = RED_F
            elif 5 <= i <= 8: c = "1F4E79"  # dark blue for source refs
            _cell(ws_a.cell(r, i+1), v, bg=bg, color=c,
                  bold=(i in (5, 6)),  # bold for file and function
                  align="center" if i < 5 or i in (7, 8) else "left")

    for i, w in enumerate([6, 20, 10, 28, 14, 30, 24, 8, 12, 55, 40, 40]):
        ws_a.column_dimensions[get_column_letter(i+1)].width = w

    ws_a.sheet_properties.tabColor = "7030A0"

    # ── CONFIG SHEET ─────────────────────────────────────────────────────────
    ws_c = wb.create_sheet("CONFIG")
    ws_c.sheet_view.showGridLines = False

    ws_c.merge_cells("A1:D1")
    _hdr(ws_c["A1"], "VALIDATION CONFIGURATION — For Reproducibility", sz=11)
    ws_c.row_dimensions[1].height = 30

    r = 3
    _hdr(ws_c.cell(r, 1), "Parameter")
    _hdr(ws_c.cell(r, 2), "Value")
    for key, val in study.items():
        r += 1
        _cell(ws_c.cell(r, 1), key, bg=GREY, bold=True, color=NAVY)
        _cell(ws_c.cell(r, 2), str(val))

    r += 2
    _hdr(ws_c.cell(r, 1), "Tolerance Setting")
    _hdr(ws_c.cell(r, 2), "Value")
    for key, val in config.get("tolerances", {}).items():
        r += 1
        _cell(ws_c.cell(r, 1), key, bg=GREY, bold=True, color=NAVY)
        _cell(ws_c.cell(r, 2), str(val))

    r += 2
    _hdr(ws_c.cell(r, 1), "TFL ID", bg=BLUE)
    _hdr(ws_c.cell(r, 2), "Title", bg=BLUE)
    _hdr(ws_c.cell(r, 3), "Dataset", bg=BLUE)
    _hdr(ws_c.cell(r, 4), "Population Filter", bg=BLUE)
    for tfl_cfg in config.get("tfl_configs", []):
        r += 1
        _cell(ws_c.cell(r, 1), tfl_cfg["tfl_id"], bold=True, color=NAVY)
        _cell(ws_c.cell(r, 2), tfl_cfg["title"])
        _cell(ws_c.cell(r, 3), tfl_cfg.get("dataset_name", ""))
        _cell(ws_c.cell(r, 4), tfl_cfg.get("population_filter", ""))

    ws_c.column_dimensions["A"].width = 26
    ws_c.column_dimensions["B"].width = 52
    ws_c.column_dimensions["C"].width = 14
    ws_c.column_dimensions["D"].width = 28

    # ── PROTOCOL VALIDATION SHEET ────────────────────────────────────────────
    protocol_results = config.get("protocol_results", [])
    if protocol_results:
        ws_p = wb.create_sheet("PROTOCOL")
        ws_p.sheet_view.showGridLines = False

        ws_p.merge_cells("A1:H1")
        _hdr(ws_p["A1"], "PROTOCOL CROSS-VALIDATION — TFL vs Protocol Document", sz=11)
        ws_p.row_dimensions[1].height = 30

        proto_meta = config.get("protocol_metadata")
        if proto_meta:
            ws_p.merge_cells("A2:H2")
            _cell(ws_p["A2"],
                  f"Study: {proto_meta.study_id} | Arms: {', '.join(proto_meta.arms.keys())} | "
                  f"Endpoints: {len(proto_meta.primary_endpoints)} primary, {len(proto_meta.secondary_endpoints)} secondary",
                  bg=LBLUE, bold=True, color=NAVY, align="center")

        r = 4
        proto_hdrs = ["TFL ID", "Category", "Check", "Protocol Value", "TFL Value", "Result", "Note"]
        for i, h in enumerate(proto_hdrs):
            _hdr(ws_p.cell(r, i+1), h)
        ws_p.row_dimensions[r].height = 26

        for pvr in protocol_results:
            for cv in pvr.cross_validations:
                r += 1
                match = cv["match"]
                result_bg = GREEN if match else RED
                result_color = GREEN_F if match else RED_F
                row_bg = WHITE if match else "FFF0F0"

                _cell(ws_p.cell(r, 1), pvr.tfl_id, bg=row_bg, bold=True, color=NAVY)
                _cell(ws_p.cell(r, 2), pvr.category, bg=row_bg)
                _cell(ws_p.cell(r, 3), cv["check"], bg=row_bg, bold=True)
                _cell(ws_p.cell(r, 4), cv["protocol_value"], bg=row_bg)
                _cell(ws_p.cell(r, 5), cv["tfl_value"], bg=row_bg)
                _cell(ws_p.cell(r, 6), "PASS" if match else "FAIL", bg=result_bg, bold=True, color=result_color, align="center")
                _cell(ws_p.cell(r, 7), cv["note"], bg=row_bg)

            # Add warnings
            for w in pvr.warnings:
                r += 1
                _cell(ws_p.cell(r, 1), pvr.tfl_id, bg=YELLOW, bold=True, color=NAVY)
                _cell(ws_p.cell(r, 2), pvr.category, bg=YELLOW)
                _cell(ws_p.cell(r, 3), "WARNING", bg=YELLOW, bold=True, color="7F6000")
                ws_p.merge_cells(f"D{r}:G{r}")
                _cell(ws_p.cell(r, 4), w, bg=YELLOW, color="7F6000")

        ws_p.column_dimensions["A"].width = 10
        ws_p.column_dimensions["B"].width = 16
        ws_p.column_dimensions["C"].width = 30
        ws_p.column_dimensions["D"].width = 40
        ws_p.column_dimensions["E"].width = 40
        ws_p.column_dimensions["F"].width = 10
        ws_p.column_dimensions["G"].width = 50
        ws_p.sheet_properties.tabColor = "2E75B6"

    # ── SAP VALIDATION SHEET ─────────────────────────────────────────────────
    sap_results = config.get("sap_results", [])
    if sap_results:
        ws_s2 = wb.create_sheet("SAP")
        ws_s2.sheet_view.showGridLines = False

        ws_s2.merge_cells("A1:H1")
        _hdr(ws_s2["A1"], "SAP CROSS-VALIDATION — TFL vs Statistical Analysis Plan", sz=11)
        ws_s2.row_dimensions[1].height = 30

        sap_meta = config.get("sap_metadata")
        if sap_meta:
            ws_s2.merge_cells("A2:H2")
            _cell(ws_s2["A2"],
                  f"SAP Version: {sap_meta.version} | Methods: {', '.join(list(sap_meta.statistical_methods.keys())[:4])} | "
                  f"Analyses: {len(sap_meta.primary_analyses)} primary, {len(sap_meta.secondary_analyses)} secondary",
                  bg=LBLUE, bold=True, color=NAVY, align="center")

        r = 4
        sap_hdrs = ["TFL ID", "Category", "Check", "SAP Specification", "TFL Value", "Result", "Note"]
        for i, h in enumerate(sap_hdrs):
            _hdr(ws_s2.cell(r, i+1), h)
        ws_s2.row_dimensions[r].height = 26

        for svr in sap_results:
            for cv in svr.cross_validations:
                r += 1
                match = cv["match"]
                result_bg = GREEN if match else RED
                result_color = GREEN_F if match else RED_F
                row_bg = WHITE if match else "FFF0F0"

                _cell(ws_s2.cell(r, 1), svr.tfl_id, bg=row_bg, bold=True, color=NAVY)
                _cell(ws_s2.cell(r, 2), svr.category, bg=row_bg)
                _cell(ws_s2.cell(r, 3), cv["check"], bg=row_bg, bold=True)
                _cell(ws_s2.cell(r, 4), cv["sap_specification"], bg=row_bg)
                _cell(ws_s2.cell(r, 5), cv["tfl_value"], bg=row_bg)
                _cell(ws_s2.cell(r, 6), "PASS" if match else "FAIL", bg=result_bg, bold=True, color=result_color, align="center")
                _cell(ws_s2.cell(r, 7), cv["note"], bg=row_bg)

            for w in svr.warnings:
                r += 1
                _cell(ws_s2.cell(r, 1), svr.tfl_id, bg=YELLOW, bold=True, color=NAVY)
                _cell(ws_s2.cell(r, 2), svr.category, bg=YELLOW)
                _cell(ws_s2.cell(r, 3), "WARNING", bg=YELLOW, bold=True, color="7F6000")
                ws_s2.merge_cells(f"D{r}:G{r}")
                _cell(ws_s2.cell(r, 4), w, bg=YELLOW, color="7F6000")

        ws_s2.column_dimensions["A"].width = 10
        ws_s2.column_dimensions["B"].width = 16
        ws_s2.column_dimensions["C"].width = 30
        ws_s2.column_dimensions["D"].width = 40
        ws_s2.column_dimensions["E"].width = 40
        ws_s2.column_dimensions["F"].width = 10
        ws_s2.column_dimensions["G"].width = 50
        ws_s2.sheet_properties.tabColor = "ED7D31"

    # ── SAS PROGRAMS MANIFEST SHEET ──────────────────────────────────────────
    sas_programs = config.get("sas_programs", [])
    if sas_programs:
        ws_sas = wb.create_sheet("SAS PROGRAMS")
        ws_sas.sheet_view.showGridLines = False

        ws_sas.merge_cells("A1:F1")
        _hdr(ws_sas["A1"], "SAS VALIDATION PROGRAMS — Generated Code Manifest", sz=11)
        ws_sas.row_dimensions[1].height = 30

        ws_sas.merge_cells("A2:F2")
        sas_dir = config.get("validation_options", {}).get("sas_output_dir", "generated_sas")
        _cell(ws_sas["A2"],
              f"Output directory: {sas_dir} | Total programs: {len(sas_programs)}",
              bg=LBLUE, bold=True, color=NAVY, align="center")

        r = 4
        sas_hdrs = ["TFL ID", "Title", "Category", "Filename", "Lines", "Status"]
        for i, h in enumerate(sas_hdrs):
            _hdr(ws_sas.cell(r, i+1), h)
        ws_sas.row_dimensions[r].height = 26

        for prog in sas_programs:
            r += 1
            lines = len(prog.full_code.splitlines())
            _cell(ws_sas.cell(r, 1), prog.tfl_id, bold=True, color=NAVY)
            _cell(ws_sas.cell(r, 2), prog.tfl_title)
            _cell(ws_sas.cell(r, 3), prog.category, align="center")
            _cell(ws_sas.cell(r, 4), prog.filename, color=BLUE)
            _cell(ws_sas.cell(r, 5), lines, align="center")
            _cell(ws_sas.cell(r, 6), "Generated", bg=GREEN, bold=True, color=GREEN_F, align="center")

        ws_sas.column_dimensions["A"].width = 10
        ws_sas.column_dimensions["B"].width = 50
        ws_sas.column_dimensions["C"].width = 16
        ws_sas.column_dimensions["D"].width = 30
        ws_sas.column_dimensions["E"].width = 10
        ws_sas.column_dimensions["F"].width = 14
        ws_sas.sheet_properties.tabColor = "375623"

    # ── ADAM SPECS SHEET ─────────────────────────────────────────────────────
    adam_specs = config.get("adam_specs")
    if adam_specs and adam_specs.get("datasets"):
        ws_s = wb.create_sheet("ADAM SPECS")
        ws_s.sheet_view.showGridLines = False

        ws_s.merge_cells("A1:H1")
        _hdr(ws_s["A1"], "ADaM SPECIFICATIONS — Variable Definitions from Specs File", sz=11)
        ws_s.row_dimensions[1].height = 30

        ws_s.merge_cells("A2:H2")
        specs_file = adam_specs.get("filepath", "")
        _cell(ws_s["A2"],
              f"Source: {os.path.basename(specs_file) if specs_file else 'N/A'} — "
              "Variable metadata used to drive validation and logged for audit traceability",
              bg=LBLUE, bold=True, color=NAVY, align="center")

        r = 4
        for ds_name, variables in sorted(adam_specs["datasets"].items()):
            # Dataset header
            ws_s.merge_cells(f"A{r}:H{r}")
            _cell(ws_s.cell(r, 1), f"Dataset: {ds_name}", bg=NAVY, bold=True, color="FFFFFF")
            ws_s.cell(r, 1).font = Font(name="Arial", size=11, bold=True, color=WHITE)
            r += 1

            # Column headers
            spec_headers = ["Variable", "Label", "Type", "Length", "Format",
                            "Controlled Terms", "Derivation / Source", "Core"]
            for i, h in enumerate(spec_headers):
                _hdr(ws_s.cell(r, i+1), h, bg=BLUE)
            r += 1

            for var_name in sorted(variables.keys()):
                meta = variables[var_name]
                vals = [
                    meta.get("variable", var_name),
                    meta.get("label", ""),
                    meta.get("type", ""),
                    meta.get("length", ""),
                    meta.get("format", ""),
                    meta.get("codelist", ""),
                    meta.get("derivation", ""),
                    meta.get("core", ""),
                ]
                is_key = meta.get("core", "").lower() == "req"
                for i, v in enumerate(vals):
                    bg = GREY if is_key else WHITE
                    _cell(ws_s.cell(r, i+1), v, bg=bg,
                          bold=(i == 0),
                          color=NAVY if i == 0 else "000000")
                r += 1
            r += 1  # blank row between datasets

        ws_s.column_dimensions["A"].width = 16
        ws_s.column_dimensions["B"].width = 38
        ws_s.column_dimensions["C"].width = 8
        ws_s.column_dimensions["D"].width = 8
        ws_s.column_dimensions["E"].width = 12
        ws_s.column_dimensions["F"].width = 35
        ws_s.column_dimensions["G"].width = 65
        ws_s.column_dimensions["H"].width = 8
        ws_s.sheet_properties.tabColor = "2E75B6"

    # ── Save ─────────────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"Validation report saved to: {output_path}")
    return output_path
